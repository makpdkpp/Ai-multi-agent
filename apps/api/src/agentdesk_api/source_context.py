from __future__ import annotations

import csv
from datetime import UTC, datetime
from io import BytesIO, StringIO
from pathlib import PurePosixPath
from typing import Any

import boto3
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from agentdesk_api.config import Settings
from agentdesk_api.db.models import Agent, AgentDataSource, DataSource, SourceFile

MAX_CONTEXT_CHARS = 18000
MAX_ROWS_PER_SHEET = 80
MAX_COLUMNS_PER_SHEET = 30
MAX_ATTACHED_SOURCES = 5


def _s3_client(settings: Settings):
    return boto3.client(
        "s3",
        endpoint_url=settings.s3_endpoint_url,
        aws_access_key_id=settings.s3_access_key,
        aws_secret_access_key=settings.s3_secret_key,
    )


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.date().isoformat()
        return value.astimezone(UTC).isoformat()
    return str(value).strip()


def _headers(first_row: list[Any]) -> list[str]:
    return [
        _cell_text(value) or f"Column {index + 1}"
        for index, value in enumerate(first_row[:MAX_COLUMNS_PER_SHEET])
    ]


def _format_rows(headers: list[str], rows: list[list[Any]]) -> str:
    if not headers:
        return "(no columns)"
    lines = ["| " + " | ".join(headers) + " |"]
    lines.append("| " + " | ".join("---" for _ in headers) + " |")
    for row in rows[:MAX_ROWS_PER_SHEET]:
        values = [
            _cell_text(row[index]) if index < len(row) else ""
            for index in range(len(headers))
        ]
        lines.append("| " + " | ".join(value.replace("|", "\\|") for value in values) + " |")
    return "\n".join(lines)


def _csv_sheets(content: bytes) -> list[tuple[str, list[str], list[list[Any]], int]]:
    text = content.decode("utf-8-sig")
    rows = list(csv.reader(StringIO(text)))
    if not rows:
        return [("CSV", [], [], 0)]
    headers = _headers(rows[0])
    data_rows = [row[:MAX_COLUMNS_PER_SHEET] for row in rows[1:]]
    return [("CSV", headers, data_rows, len(data_rows))]


def _workbook_sheets(content: bytes) -> list[tuple[str, list[str], list[list[Any]], int]]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    try:
        sheets: list[tuple[str, list[str], list[list[Any]], int]] = []
        for worksheet in workbook.worksheets:
            row_iter = worksheet.iter_rows(values_only=True)
            first_row = list(next(row_iter, ()) or ())
            headers = _headers(first_row)
            rows: list[list[Any]] = []
            for row in row_iter:
                rows.append(list(row[:MAX_COLUMNS_PER_SHEET]))
                if len(rows) >= MAX_ROWS_PER_SHEET:
                    break
            total_rows = max(0, (worksheet.max_row or 0) - 1)
            sheets.append((worksheet.title, headers, rows, total_rows))
        return sheets
    finally:
        workbook.close()


def _excel_sheets(
    content: bytes,
    filename: str,
) -> list[tuple[str, list[str], list[list[Any]], int]]:
    if PurePosixPath(filename.lower()).suffix == ".csv":
        return _csv_sheets(content)
    return _workbook_sheets(content)


def _latest_ready_file(source: DataSource) -> SourceFile | None:
    ready_files = [source_file for source_file in source.files if source_file.status == "ready"]
    if not ready_files:
        return None
    return max(ready_files, key=lambda item: (item.version, item.created_at))


def _fallback_preview(source: DataSource, source_file: SourceFile) -> str:
    sheets = source_file.file_metadata.get("sheets", [])
    lines = [
        f"### Source: {source.name}",
        f"- Type: {source.source_type}",
        f"- File: {source_file.original_name}",
        "- Note: full file could not be loaded; using stored preview metadata.",
    ]
    for sheet in sheets if isinstance(sheets, list) else []:
        if not isinstance(sheet, dict):
            continue
        name = sheet.get("name", "Sheet")
        columns = sheet.get("columns", [])
        preview_rows = sheet.get("preview_rows", [])
        lines.append(f"\nSheet: {name}")
        lines.append(f"Columns: {', '.join(str(column) for column in columns)}")
        if preview_rows:
            lines.append(f"Preview rows: {preview_rows}")
    return "\n".join(lines)


def _source_block(source: DataSource, source_file: SourceFile, content: bytes) -> str:
    lines = [
        f"### Source: {source.name}",
        f"- Type: {source.source_type}",
        f"- File: {source_file.original_name}",
        f"- Loaded rows per sheet limit: {MAX_ROWS_PER_SHEET}",
    ]
    for sheet_name, headers, rows, total_rows in _excel_sheets(content, source_file.original_name):
        lines.append(f"\nSheet: {sheet_name} ({total_rows} data rows)")
        lines.append(_format_rows(headers, rows))
    return "\n".join(lines)


async def build_agent_data_source_context(
    session: AsyncSession,
    settings: Settings,
    agent: Agent,
) -> str:
    result = await session.execute(
        select(AgentDataSource)
        .options(selectinload(AgentDataSource.data_source).selectinload(DataSource.files))
        .where(
            AgentDataSource.agent_id == agent.id,
            AgentDataSource.department_id == agent.department_id,
            AgentDataSource.enabled.is_(True),
        )
        .order_by(AgentDataSource.priority, AgentDataSource.created_at)
        .limit(MAX_ATTACHED_SOURCES)
    )
    links = list(result.scalars().all())
    blocks: list[str] = []
    client = None
    for link in links:
        source = link.data_source
        if source is None or source.source_type != "excel" or source.status != "ready":
            continue
        source_file = _latest_ready_file(source)
        if source_file is None:
            continue
        try:
            client = client or _s3_client(settings)
            response = client.get_object(Bucket=settings.s3_bucket, Key=source_file.object_key)
            content = response["Body"].read()
            block = _source_block(source, source_file, content)
        except Exception:
            block = _fallback_preview(source, source_file)
        blocks.append(block)

    if not blocks:
        return ""

    context = "\n\n".join(blocks)
    if len(context) > MAX_CONTEXT_CHARS:
        context = f"{context[:MAX_CONTEXT_CHARS]}\n...[data source context truncated]"
    return (
        "You have access to the following attached Excel data sources for this agent. "
        "Use only this data when answering questions about uploaded files or company records. "
        "If the answer is not present in this context, say that the attached data does "
        "not contain it.\n\n"
        f"{context}"
    )
