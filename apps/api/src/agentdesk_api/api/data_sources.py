from __future__ import annotations

import csv
import hashlib
from datetime import UTC, datetime
from io import BytesIO, StringIO
from pathlib import PurePosixPath
from typing import Annotated
from uuid import UUID, uuid4

import boto3
from fastapi import APIRouter, File, Form, HTTPException, UploadFile, status
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import selectinload

from agentdesk_api.api.agents import (
    get_agent_or_404,
    require_agent_member_access,
    require_department_agent_manager_access,
    require_department_member_access,
    set_agent_department_context,
    set_auth_context,
)
from agentdesk_api.api.auth import AppSettings, AuthDependency, CsrfDependency, DbSession
from agentdesk_api.db.models import AgentDataSource, DataSource, SourceChunk, SourceFile

router = APIRouter(tags=["data-sources"])

EXCEL_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroEnabled.12",
    "text/csv",
    "application/csv",
}
EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}
MAX_EXCEL_UPLOAD_BYTES = 20 * 1024 * 1024
MAX_PREVIEW_ROWS = 5
MAX_PREVIEW_COLUMNS = 30


class AgentDataSourceAttach(BaseModel):
    data_source_id: UUID
    access_scope: str = Field(default="internal_only", pattern="^(internal_only|public_allowed)$")
    priority: int = Field(default=100, ge=1, le=1000)
    enabled: bool = True


def s3_client(settings: AppSettings):
    return boto3.client(
        "s3",
        endpoint_url=settings.s3_endpoint_url,
        aws_access_key_id=settings.s3_access_key,
        aws_secret_access_key=settings.s3_secret_key,
    )


def safe_filename(filename: str) -> str:
    name = PurePosixPath(filename.replace("\\", "/")).name
    return name or "workbook.xlsx"


def validate_excel_upload(filename: str, content_type: str | None, size: int) -> str:
    extension = PurePosixPath(filename.lower()).suffix
    if extension not in EXCEL_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="รองรับเฉพาะไฟล์ .xlsx, .xlsm และ .csv สำหรับ Excel source",
        )
    if content_type and content_type not in EXCEL_MIME_TYPES and extension != ".csv":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="ชนิดไฟล์ไม่อยู่ใน allowlist สำหรับ Excel source",
        )
    if size > MAX_EXCEL_UPLOAD_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="ไฟล์ Excel ต้องมีขนาดไม่เกิน 20 MB สำหรับ pilot",
        )
    return extension


def normalize_cell(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def parse_csv_preview(content: bytes) -> dict[str, object]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(StringIO(text))
    rows = list(reader)
    headers = [
        str(item).strip() or f"Column {index + 1}"
        for index, item in enumerate(rows[0] if rows else [])
    ]
    preview_rows = rows[1 : MAX_PREVIEW_ROWS + 1]
    return {
        "sheets": [
            {
                "name": "CSV",
                "row_count": max(0, len(rows) - 1),
                "column_count": len(headers),
                "columns": headers[:MAX_PREVIEW_COLUMNS],
                "preview_rows": [
                    {
                        headers[index] if index < len(headers) else f"Column {index + 1}": value
                        for index, value in enumerate(row[:MAX_PREVIEW_COLUMNS])
                    }
                    for row in preview_rows
                ],
            }
        ]
    }


def parse_workbook_preview(content: bytes) -> dict[str, object]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    sheets: list[dict[str, object]] = []
    for worksheet in workbook.worksheets:
        row_iter = worksheet.iter_rows(values_only=True)
        first_row = next(row_iter, None)
        headers = [
            str(value).strip()
            if value is not None and str(value).strip()
            else f"Column {index + 1}"
            for index, value in enumerate(first_row or [])
        ]
        preview_rows = []
        for row in row_iter:
            preview_rows.append(
                {
                    (
                        headers[index] if index < len(headers) else f"Column {index + 1}"
                    ): normalize_cell(value)
                    for index, value in enumerate(row[:MAX_PREVIEW_COLUMNS])
                }
            )
            if len(preview_rows) >= MAX_PREVIEW_ROWS:
                break
        sheets.append(
            {
                "name": worksheet.title,
                "row_count": max(0, (worksheet.max_row or 0) - 1),
                "column_count": worksheet.max_column or len(headers),
                "columns": headers[:MAX_PREVIEW_COLUMNS],
                "preview_rows": preview_rows,
            }
        )
    workbook.close()
    return {"sheets": sheets}


def parse_excel_preview(content: bytes, extension: str) -> dict[str, object]:
    if extension == ".csv":
        return parse_csv_preview(content)
    return parse_workbook_preview(content)


def build_index_chunks(
    content: bytes, extension: str, max_rows: int = 5000
) -> list[dict[str, object]]:
    """Create row chunks for retrieval without sending the full workbook to the LLM."""
    if extension == ".csv":
        rows = list(csv.reader(StringIO(content.decode("utf-8-sig"))))
        sheets = [("CSV", rows)]
    else:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        sheets = [
            (worksheet.title, list(worksheet.iter_rows(values_only=True)))
            for worksheet in workbook.worksheets
        ]
        workbook.close()
    chunks: list[dict[str, object]] = []
    for sheet_name, rows in sheets:
        if not rows:
            continue
        headers = [
            str(value).strip() if value is not None and str(value).strip() else f"Column {i + 1}"
            for i, value in enumerate(rows[0])
        ]
        for row_number, row in enumerate(rows[1 : max_rows + 1], start=2):
            values = [normalize_cell(value) for value in row[:MAX_PREVIEW_COLUMNS]]
            pairs = [
                f"{headers[i]}: {values[i]}"
                for i in range(min(len(headers), len(values)))
                if values[i] not in (None, "")
            ]
            if pairs:
                chunks.append(
                    {
                        "content": f"Sheet: {sheet_name}; Row: {row_number}; " + "; ".join(pairs),
                        "metadata": {"sheet": sheet_name, "row": row_number},
                    }
                )
    return chunks


def source_file_data(source_file: SourceFile) -> dict[str, object]:
    return {
        "id": str(source_file.id),
        "data_source_id": str(source_file.data_source_id),
        "original_name": source_file.original_name,
        "mime_type": source_file.mime_type,
        "size_bytes": source_file.size_bytes,
        "sha256": source_file.sha256,
        "status": source_file.status,
        "version": source_file.version,
        "metadata": source_file.file_metadata,
        "processing_error": source_file.processing_error,
        "created_at": source_file.created_at.isoformat(),
        "indexed_at": source_file.indexed_at.isoformat() if source_file.indexed_at else None,
    }


def data_source_data(data_source: DataSource) -> dict[str, object]:
    return {
        "id": str(data_source.id),
        "department_id": str(data_source.department_id),
        "name": data_source.name,
        "source_type": data_source.source_type,
        "status": data_source.status,
        "connection_config": data_source.connection_config,
        "allowed_schema": data_source.allowed_schema,
        "files": [source_file_data(source_file) for source_file in data_source.files],
        "created_at": data_source.created_at.isoformat(),
        "updated_at": data_source.updated_at.isoformat(),
    }


def agent_source_data(link: AgentDataSource) -> dict[str, object]:
    return {
        "id": str(link.id),
        "department_id": str(link.department_id),
        "agent_id": str(link.agent_id),
        "data_source_id": str(link.data_source_id),
        "data_source_name": link.data_source.name if link.data_source else None,
        "source_type": link.data_source.source_type if link.data_source else None,
        "status": link.data_source.status if link.data_source else None,
        "access_scope": link.access_scope,
        "priority": link.priority,
        "enabled": link.enabled,
        "created_at": link.created_at.isoformat(),
    }


async def get_data_source_or_404(data_source_id: UUID, session: DbSession) -> DataSource:
    data_source = await session.scalar(
        select(DataSource)
        .options(selectinload(DataSource.files))
        .where(DataSource.id == data_source_id, DataSource.deleted_at.is_(None))
    )
    if data_source is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Data source not found.")
    return data_source


@router.get("/departments/{department_id}/data-sources")
async def list_department_data_sources(
    department_id: UUID,
    auth: AuthDependency,
    session: DbSession,
) -> dict[str, object]:
    await require_department_member_access(department_id, auth, session)
    result = await session.execute(
        select(DataSource)
        .options(selectinload(DataSource.files))
        .where(DataSource.department_id == department_id, DataSource.deleted_at.is_(None))
        .order_by(DataSource.updated_at.desc(), DataSource.created_at.desc())
    )
    sources = list(result.scalars().all())
    return {
        "data": [data_source_data(source) for source in sources],
        "meta": {"total": len(sources)},
    }


@router.post("/departments/{department_id}/data-sources/excel", status_code=status.HTTP_201_CREATED)
async def upload_excel_data_source(
    department_id: UUID,
    auth: AuthDependency,
    _: CsrfDependency,
    session: DbSession,
    settings: AppSettings,
    name: Annotated[str, Form(min_length=1, max_length=200)],
    file: Annotated[UploadFile, File()],
) -> dict[str, object]:
    await require_department_agent_manager_access(department_id, auth, session)
    content = await file.read()
    original_name = safe_filename(file.filename or "workbook.xlsx")
    extension = validate_excel_upload(original_name, file.content_type, len(content))
    file_hash = hashlib.sha256(content).hexdigest()
    object_key = f"departments/{department_id}/excel/{uuid4()}/{original_name}"

    metadata: dict[str, object]
    status_value = "ready"
    processing_error = None
    try:
        metadata = parse_excel_preview(content, extension)
    except Exception:
        metadata = {"sheets": []}
        status_value = "error"
        processing_error = "ไม่สามารถอ่านโครงสร้างไฟล์ Excel ได้"

    s3_client(settings).put_object(
        Bucket=settings.s3_bucket,
        Key=object_key,
        Body=content,
        ContentType=file.content_type or "application/octet-stream",
        Metadata={"sha256": file_hash},
    )

    data_source = DataSource(
        department_id=department_id,
        name=name.strip(),
        source_type="excel",
        status=status_value,
        connection_config={"storage": "minio", "latest_object_key": object_key},
        allowed_schema=metadata,
        created_by=auth.user_id,
        updated_by=auth.user_id,
    )
    source_file = SourceFile(
        department_id=department_id,
        data_source=data_source,
        object_key=object_key,
        original_name=original_name,
        mime_type=file.content_type or "application/octet-stream",
        size_bytes=len(content),
        sha256=file_hash,
        status="ready" if status_value == "ready" else "failed",
        version=1,
        file_metadata=metadata,
        processing_error=processing_error,
        uploaded_by=auth.user_id,
        indexed_at=datetime.now(UTC) if status_value == "ready" else None,
    )
    session.add_all([data_source, source_file])
    await session.flush()
    if status_value == "ready":
        for index, chunk in enumerate(build_index_chunks(content, extension)):
            session.add(
                SourceChunk(
                    department_id=department_id,
                    data_source_id=data_source.id,
                    source_file_id=source_file.id,
                    chunk_index=index,
                    content=str(chunk["content"]),
                    metadata=chunk["metadata"],
                )
            )
        data_source.status = "ready"
    await session.refresh(data_source, attribute_names=["files"])
    response_data = data_source_data(data_source)
    await session.commit()
    return {"data": response_data}


@router.get("/data-sources/{data_source_id}")
async def get_data_source(
    data_source_id: UUID,
    auth: AuthDependency,
    session: DbSession,
) -> dict[str, object]:
    await set_auth_context(session, auth)
    data_source = await get_data_source_or_404(data_source_id, session)
    await require_department_member_access(data_source.department_id, auth, session)
    return {"data": data_source_data(data_source)}


@router.get("/agents/{agent_id}/data-sources")
async def list_agent_data_sources(
    agent_id: UUID,
    auth: AuthDependency,
    session: DbSession,
) -> dict[str, object]:
    await set_auth_context(session, auth)
    agent = await get_agent_or_404(agent_id, session)
    await require_agent_member_access(agent, auth, session)
    result = await session.execute(
        select(AgentDataSource)
        .options(selectinload(AgentDataSource.data_source))
        .where(AgentDataSource.agent_id == agent.id)
        .order_by(AgentDataSource.priority, AgentDataSource.created_at)
    )
    links = list(result.scalars().all())
    return {"data": [agent_source_data(link) for link in links], "meta": {"total": len(links)}}


@router.post("/agents/{agent_id}/data-sources", status_code=status.HTTP_201_CREATED)
async def attach_agent_data_source(
    agent_id: UUID,
    payload: AgentDataSourceAttach,
    auth: AuthDependency,
    _: CsrfDependency,
    session: DbSession,
) -> dict[str, object]:
    await set_auth_context(session, auth)
    agent = await get_agent_or_404(agent_id, session)
    await require_department_agent_manager_access(agent.department_id, auth, session)
    data_source = await get_data_source_or_404(payload.data_source_id, session)
    if data_source.department_id != agent.department_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Data source not found.")
    await set_agent_department_context(session, auth, agent)
    link = AgentDataSource(
        department_id=agent.department_id,
        agent_id=agent.id,
        data_source_id=data_source.id,
        access_scope=payload.access_scope,
        priority=payload.priority,
        enabled=payload.enabled,
    )
    session.add(link)
    try:
        await session.flush()
        await session.refresh(link, attribute_names=["data_source"])
        response_data = agent_source_data(link)
        await session.commit()
    except IntegrityError as exc:
        await session.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Data source is already attached to this agent.",
        ) from exc
    return {"data": response_data}


@router.delete(
    "/agents/{agent_id}/data-sources/{data_source_id}",
    status_code=status.HTTP_204_NO_CONTENT,
)
async def detach_agent_data_source(
    agent_id: UUID,
    data_source_id: UUID,
    auth: AuthDependency,
    _: CsrfDependency,
    session: DbSession,
) -> None:
    await set_auth_context(session, auth)
    agent = await get_agent_or_404(agent_id, session)
    await require_department_agent_manager_access(agent.department_id, auth, session)
    link = await session.scalar(
        select(AgentDataSource).where(
            AgentDataSource.agent_id == agent.id,
            AgentDataSource.data_source_id == data_source_id,
        )
    )
    if link is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Agent data source not found.",
        )
    await session.delete(link)
    await session.commit()
